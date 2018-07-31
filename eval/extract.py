from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cifs
import re


def find_start(sheet):
    for row in sheet.rows:
        for cell in row:
            if (unicode(cell.value) or '').strip() == u'Term ID':
                print 'start {}'.format(cell.coordinate)
                yield sheet.cell(row=cell.row, column=cifs(cell.column))
                break


def find_end(sheet, start):
    cell_ = start
    end_col = end_row = None
    c = cifs(cell_.column) + 1
    while True:
        cell = sheet.cell(row=cell_.row, column=c)
        if (unicode(cell.value) or '').strip().lower() == u'frequency':
            end_col = c
            break
        c += 1
    r = cell_.row + 1
    while True:
        cell = sheet.cell(row=r, column=cifs(cell_.column))
        text = unicode(cell.value or '').strip()
        if u"HP" not in text.upper() and len(text) > 0 and not text.lower() == 'update':
            end_row = r - 1
            break
        r += 1
    if (end_col or end_row) is None:
        raise Exception("No ending cell found for {}".format(cell_.coordinate))
    end_cell = sheet.cell(row=end_row, column=end_col)
    print 'end {}'.format(end_cell.coordinate)
    return end_cell


def do(filename):
    wb = load_workbook(filename)
    sheet = wb._sheets[0]

    for start in find_start(sheet):
        end = find_end(sheet, start)
        cell_range = sheet[start.coordinate : end.coordinate]
        yield cell_range


def store_editable(filename, out_file):
    wb = Workbook()

    si = 0
    for _range in do(filename):
        sheet = wb.create_sheet("S{:>02}".format(si))
        for i in xrange(len(_range)):
            row = _range[i]
            for j in xrange(len(row)):
                cell = row[j]
                sheet.cell(row=i+1, column=j+1).value = cell.value
        si += 1

    wb.save(out_file)


def merge_cells(sheet, _range):
    start_coor, end_coor = _range.split(':')
    cell_range = sheet[start_coor : end_coor]
    cells = reduce(lambda x,y: list(x) + list(y), cell_range, [])
    coords = map(lambda c: c.coordinate, cells)
    coords.pop(0)
    cell_values = map(lambda c: (c.value or ''), cells)
    merged_value = reduce(lambda x,y: str(x)+str(y), cell_values, '')
    sheet[start_coor].value = merged_value
    for coor in coords:
        sheet[coor].value = None


def process_range(sheet):
    # merge columns
    def find_max_col_to_merge():
        titleRow = sheet.rows.next()
        for cell in titleRow:
            if (cell.value or '').strip().lower() == 'frequency':
                return cifs(cell.column) - 1
    max_col_to_merge = find_max_col_to_merge()
    def find_rows_to_merge():
        freq_col = max_col_to_merge + 1
        for ri in xrange(sheet.max_row, 2, -1):
            text = str(sheet.cell(row=ri, column=freq_col).value or '').strip()
            if len(text) == 0:
                yield ri
    for ri in xrange(2, sheet.max_row+1):
        cell_start = sheet.cell(row=ri, column=1)
        cell_end = sheet.cell(row=ri, column=max_col_to_merge)
        merge_cells(sheet, "{}:{}".format(cell_start.coordinate, cell_end.coordinate))
    # merge rows with no frequency
    for ri in find_rows_to_merge():
        cell_start = sheet.cell(row=ri-1, column=1)
        cell_end = sheet.cell(row=ri, column=max_col_to_merge)
        merge_cells(sheet, "{}:{}".format(cell_start.coordinate, cell_end.coordinate))
    # clear sheet, keeping only first 2 columns
    sheet.delete_cols(idx=2, amount=max_col_to_merge+1-2)
    # clear empty rows
    for ri in xrange(sheet.max_row, 1, -1):
        text = str(sheet.cell(row=ri, column=1).value or '').strip()
        text += str(sheet.cell(row=ri, column=2).value or '').strip()
        if len(text) == 0:
            sheet.delete_rows(idx=ri)
    for row in sheet.rows:
        for cell in row:
            if not (isinstance(cell.value, unicode) or isinstance(cell.value, str)):
                continue
            cell.value = cell.value.replace("HPO", "HP")
            res = re.findall(r'(HP:([0-9]+))', cell.value)
            for t in res:
                cell.value = cell.value.replace(t[0], "HP:{:>06}".format(int(t[1])))


def main():
    store_editable("sup_material.xlsx", "out.xlsx")

    wb = load_workbook("out.xlsx")
    for ws in wb._sheets[1:]:
        print ws
        process_range(ws)
    wb.save("out_.xlsx")


if __name__ == "__main__":
    main()
