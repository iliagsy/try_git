import json
from openpyxl import load_workbook
from pymongo import MongoClient
import re


with open("data/freqMap.json") as fh:
    freqMap = json.loads(fh.read())


def _alter_hpo(hpo):
    res = re.findall(r'HP:(\d+)', hpo)
    assert len(res) == 1
    return 'HP:{:0>7}'.format(int(res[0]))


def store(sheet, db):
    gen = sheet.rows
    _ = gen.next()
    for row in gen:
        hpo, freq = map(lambda c: c.value, row[:2])
        hpo = str(hpo)
        res = re.findall(r'(HP:[0-9]{6})', hpo)
        if not len(res) > 0:
            continue
        if len(res) == 1:
            hpo = res[0]
            hpo = _alter_hpo(hpo)
        else:
            for sep in ['or', 'and', 'and/or']:
                if hpo.count(sep) + 1 == len(res):
                    res = map(_alter_hpo, res)
                    hpo = {sep: res}
                    break
        try:
            freq = float(freq)
        except:
            pass
        if not isinstance(freq, float):
            freq = str(freq).strip().lower()
            res = re.findall(r"([0-9]+)/([0-9]+)", freq)
            if len(res) > 0:
                en, dn = res[0]
                freq = float(en) / float(dn)
                freq = round(freq, 2)
            else:
                print freq
                assert freq in freqMap
                freq = freqMap[freq]
        assert isinstance(freq, float)
        db['benchmark'].update({"sid": sheet.title}, {
            "$push": {
                "hpo-freq": {
                    "hpo": hpo,
                    "frequency": freq
                }
            }
        }, upsert=True)


def store_knowledge_disease(sheet, db):
    for row in sheet.rows:
        sid, name, id_ = map(lambda c: str(c.value or '').strip(), row[:3])
        name = name.replace("\n", '')
        if id_ != '':
            db_name, db_id = id_.split()
            assert db_name == 'MIM'
            db_id = int(db_id)
            id_ = 'OMIM:{}'.format(db_id)
        else:
            id_ = name
        db['benchmark'].update({"sid": sid}, {
            "$set": {
                "disease": id_,
                "name": name
            }
        })



if __name__ == '__main__':
    db = MongoClient("localhost", 27017)['phenomizer']
    wb = load_workbook("data/out_.xlsx")
    for ws in wb._sheets[1:]:
        store(ws, db)
    
    wb = load_workbook("data/diseaseName.xlsx")
    sheet = wb.active
    store_knowledge_disease(sheet, db)
